from __future__ import annotations

import csv
import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from .link_classification import UNKNOWN, classify_link
from .link_extraction import extract_links as _extract_links

_GRAD_YEAR_RE = re.compile(r"(\d{4})\s*届")

# This is the fixed export layout used by the campus recruitment sheet.
_UPDATED_AT = 0
_COMPANY_NAME = 1
_COMPANY_NATURE = 2
_INDUSTRY = 3
_JOB_DIRECTIONS = 4
_LOCATIONS = 5
_DEADLINE = 6
_GRADUATION = 7
_EDUCATION = 8
_BATCH = 9
_ANNOUNCEMENT_SOURCE = 10
_ANNOUNCEMENT_URL = 11
_APPLY_URL = 12
_MAJOR_REQUIREMENT = 13
_HAS_WRITTEN_TEST = 14
_JUNK_COMPANIES = {"公司名称", "（必看）表格使用说明", "☝筛选前必看使用说明"}


@dataclass(frozen=True)
class SpreadsheetRow:
    row_number: int
    updated_at: datetime | None
    company_name: str | None
    company_nature: str | None
    industry: str | None
    job_directions: str | None
    locations: list[str]
    deadline_at: datetime | None
    graduation_years: list[int]
    education_requirement: str | None
    batch: str | None
    announcement_source: str | None
    announcement_url: str | None
    apply_links: list[str]
    major_requirement: str | None
    has_written_test: str | None
    source_record_id: str | None = None
    source_last_modified_at: datetime | None = None

    @property
    def recruitment_type(self) -> str | None:
        if not self.batch:
            return None
        if "实习" in self.batch:
            return "实习"
        if any(word in self.batch for word in ("秋招", "春招", "校园招聘")):
            return "校招"
        return self.batch


def clean(value: object) -> str:
    return str(value or "").strip().strip("\ufeff")


def valid_text(value: object) -> str | None:
    text = clean(value)
    return text if text and text != "/" else None


def valid_url(value: object) -> str | None:
    text = clean(value)
    return text if text.startswith(("http://", "https://")) else None


def extract_links(value: object, hyperlink_target: str | None = None) -> list[str]:
    return [
        link for link in _extract_links(value, hyperlink_target) if classify_link(link) != UNKNOWN
    ]


def parse_locations(value: object) -> list[str]:
    return [part.strip() for part in re.split(r"[,，、]", clean(value)) if part.strip()]


def parse_graduation_years(value: object) -> list[int]:
    return [int(year) for year in _GRAD_YEAR_RE.findall(clean(value))]


def parse_date(value: object) -> datetime | None:
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def extract_row(
    row_number: int, values: Sequence[object], hyperlink_target: str | None = None
) -> SpreadsheetRow | None:
    company_name = valid_text(values[_COMPANY_NAME]) if len(values) > _COMPANY_NAME else None
    directions = valid_text(values[_JOB_DIRECTIONS]) if len(values) > _JOB_DIRECTIONS else None
    if not company_name or not directions or company_name in _JUNK_COMPANIES:
        return None
    if directions in {"招聘岗位", "（必看）表格使用说明"}:
        return None

    apply_value = values[_APPLY_URL] if len(values) > _APPLY_URL else None
    return SpreadsheetRow(
        row_number=row_number,
        updated_at=parse_date(values[_UPDATED_AT] if len(values) > _UPDATED_AT else None),
        company_name=company_name,
        company_nature=valid_text(values[_COMPANY_NATURE])
        if len(values) > _COMPANY_NATURE
        else None,
        industry=valid_text(values[_INDUSTRY]) if len(values) > _INDUSTRY else None,
        job_directions=directions,
        locations=parse_locations(values[_LOCATIONS] if len(values) > _LOCATIONS else None),
        deadline_at=parse_date(values[_DEADLINE] if len(values) > _DEADLINE else None),
        graduation_years=parse_graduation_years(
            values[_GRADUATION] if len(values) > _GRADUATION else None
        ),
        education_requirement=valid_text(values[_EDUCATION] if len(values) > _EDUCATION else None),
        batch=valid_text(values[_BATCH] if len(values) > _BATCH else None),
        announcement_source=valid_text(
            values[_ANNOUNCEMENT_SOURCE] if len(values) > _ANNOUNCEMENT_SOURCE else None
        ),
        announcement_url=valid_url(
            values[_ANNOUNCEMENT_URL] if len(values) > _ANNOUNCEMENT_URL else None
        ),
        apply_links=extract_links(apply_value, hyperlink_target),
        major_requirement=valid_text(
            values[_MAJOR_REQUIREMENT] if len(values) > _MAJOR_REQUIREMENT else None
        ),
        has_written_test=valid_text(
            values[_HAS_WRITTEN_TEST] if len(values) > _HAS_WRITTEN_TEST else None
        ),
    )


def read_rows(path: Path, sheet_name: str | None = None) -> list[SpreadsheetRow]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as file:
            return [
                row
                for number, values in enumerate(csv.reader(file), start=1)
                if (row := extract_row(number, values)) is not None
            ]

    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows: list[SpreadsheetRow] = []
    for row_number in range(2, sheet.max_row + 1):
        values = [cell.value for cell in sheet[row_number]]
        link_cell = sheet.cell(row_number, _APPLY_URL + 1)
        row = extract_row(
            row_number,
            values,
            link_cell.hyperlink.target if link_cell.hyperlink else None,
        )
        if row is not None:
            rows.append(row)
    return rows


def rows_from_values(values: Iterable[Sequence[object]]) -> list[SpreadsheetRow]:
    return [
        row
        for row_number, values_row in enumerate(values, start=2)
        if (row := extract_row(row_number, values_row)) is not None
    ]


__all__ = [
    "SpreadsheetRow",
    "clean",
    "extract_links",
    "extract_row",
    "parse_date",
    "parse_graduation_years",
    "parse_locations",
    "read_rows",
    "rows_from_values",
    "valid_text",
    "valid_url",
]
