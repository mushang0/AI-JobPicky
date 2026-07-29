from __future__ import annotations

import argparse
import json
import random
from collections import defaultdict
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook

from jobpicky.collection.link_classification import UNKNOWN, classify_link
from jobpicky.collection.link_extraction import extract_links

_SEED = 20260728


def _read_records(workbook_path: Path, sheet_name: str) -> tuple[int, list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"worksheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    non_empty_cells = 0
    records: list[dict[str, object]] = []
    for row_number in range(2, sheet.max_row + 1):
        link_cell = sheet.cell(row_number, 13)
        company_cell = sheet.cell(row_number, 2)
        hyperlink_target = link_cell.hyperlink.target if link_cell.hyperlink else None
        if link_cell.value in (None, "") and not hyperlink_target:
            continue
        non_empty_cells += 1
        links = extract_links(link_cell.value, hyperlink_target)
        for link in links:
            records.append(
                {
                    "row_number": row_number,
                    "company_name": company_cell.value,
                    "raw_value": link_cell.value,
                    "url": link,
                    "link_type": classify_link(link),
                    "_links_in_cell": len(links),
                }
            )
    return non_empty_cells, records


def _choose(
    records: list[dict[str, object]], selected: list[dict[str, object]], seen: set[int]
) -> None:
    for record in records:
        key = id(record)
        if key not in seen:
            selected.append(record)
            seen.add(key)
            return


def _public_record(record: dict[str, object]) -> dict[str, object]:
    public = {
        "row_number": record["row_number"],
        "company_name": record["company_name"],
        "url": record["url"],
        "link_type": record["link_type"],
    }
    if record["raw_value"] != record["url"]:
        public["source_cell_value"] = record["raw_value"]
    return public


def _targeted_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    candidates = [record for record in records if record["link_type"] != UNKNOWN]
    query = [record for record in candidates if urlsplit(str(record["url"])).query]
    fragment = [record for record in candidates if urlsplit(str(record["url"])).fragment]
    multi = [record for record in candidates if record["_links_in_cell"] > 1]

    by_host: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in candidates:
        parts = urlsplit(str(record["url"]))
        if parts.hostname:
            by_host[parts.hostname.lower()].append(record)
    same_host_different_path: list[dict[str, object]] = []
    for host in sorted(by_host):
        group = by_host[host]
        for index, first in enumerate(group):
            for second in group[index + 1 :]:
                if urlsplit(str(first["url"])).path != urlsplit(str(second["url"])).path:
                    same_host_different_path = [first, second]
                    break
            if same_host_different_path:
                break
        if same_host_different_path:
            break
    return query[:1] + fragment[:1] + multi[:1] + same_host_different_path


def _sample_audit(
    records: list[dict[str, object]], target_size: int = 40
) -> list[dict[str, object]]:
    rng = random.Random(_SEED)
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        groups[str(record["link_type"])].append(record)
    for group in groups.values():
        rng.shuffle(group)

    selected: list[dict[str, object]] = []
    seen: set[int] = set()
    for link_type in sorted(groups):
        if link_type != UNKNOWN:
            _choose(groups[link_type], selected, seen)
    for record in groups.get(UNKNOWN, [])[: min(9, len(groups.get(UNKNOWN, [])))]:
        _choose([record], selected, seen)
    for record in _targeted_records(records):
        if len(selected) >= target_size:
            break
        _choose([record], selected, seen)

    remaining = [record for record in records if id(record) not in seen]
    rng.shuffle(remaining)
    for record in remaining:
        if len(selected) >= target_size:
            break
        unknown_count = sum(item["link_type"] == UNKNOWN for item in selected)
        if record["link_type"] == UNKNOWN and unknown_count >= 9:
            continue
        _choose([record], selected, seen)

    audit = []
    for record in selected:
        audit.append(
            _public_record(record)
            | {
                "predicted_type": record["link_type"],
                "expected_type": None,
                "review_status": "PENDING",
                "review_note": "",
            }
        )
    return audit


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Classify application links from an Excel worksheet."
    )
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--output-dir", type=Path, default=Path("artifacts/link-classification"))
    args = parser.parse_args()

    non_empty_cells, records = _read_records(args.input, args.sheet)
    results = [_public_record(record) for record in records]
    unknown = [record for record in records if record["link_type"] == UNKNOWN]
    audit = _sample_audit(records)
    _write_json(args.output_dir / "classification-results.json", results)
    _write_json(
        args.output_dir / "classification-unknown.json", [_public_record(r) for r in unknown]
    )
    _write_json(args.output_dir / "classification-audit.json", audit)

    counts: dict[str, int] = defaultdict(int)
    for record in records:
        counts[str(record["link_type"])] += 1
    summary = {"non_empty_cells": non_empty_cells, "links": len(records), "counts": counts}
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
